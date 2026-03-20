#!/usr/bin/env python3
"""
Экспорт аналитики в Excel с диаграммами.
Источники: analytics.jsonl или Excel-файл (экспорт из Google Таблицы).

Запуск:
  python scripts/export-analytics-to-excel.py
  python scripts/export-analytics-to-excel.py "C:\\path\\to\\tv.xlsx"
"""
import json
import os
import sys
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Установите: pip install openpyxl")
    sys.exit(1)

# Колонки аналитики (как в Google Sheets)
COLS = ['ts', 'page', 'event', 'sessionId', 'tgId', 'tgUsername', 'element', 'href', 'durationSec', 'title', 'lessonIndex', 'progress']


def load_from_jsonl(path):
    """Читает analytics.jsonl"""
    rows = []
    if not os.path.exists(path):
        return rows
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                row = [str(obj.get(c, ''))[:200] for c in COLS]
                rows.append(row)
            except json.JSONDecodeError:
                pass
    return rows


def load_from_excel(path):
    """Читает аналитику из Excel (лист с ts, page, event или Столбец 1,2,3)"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = []
    for ws in wb.worksheets:
        data = list(ws.iter_rows(values_only=True))
        if len(data) < 2:
            continue
        header = [str(c or '').lower() for c in data[0]]
        # Столбец 1,2,3 = ts, page, event (Looker Studio export)
        is_analytics = any(
            x in ' '.join(header) for x in ['ts', 'page', 'event', 'столбец', 'column']
        ) or (len(header) >= 3 and len(header) <= 15)
        if is_analytics:
            for r in data[1:]:
                if any(c is not None for c in r):
                    row = [str(c or '')[:200] for c in r]
                    while len(row) < len(COLS):
                        row.append('')
                    rows.append(row[:len(COLS)])
            if rows:
                break
    return rows


def main():
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    jsonl_path = os.path.join(project_root, 'analytics.jsonl')
    excel_input = sys.argv[1] if len(sys.argv) > 1 else None

    rows = []
    if excel_input and os.path.exists(excel_input):
        print(f"Читаю из Excel: {excel_input}")
        rows = load_from_excel(excel_input)
    if not rows:
        print("Читаю из analytics.jsonl")
        rows = load_from_jsonl(jsonl_path)

    if not rows:
        print("Нет данных. Укажите путь к Excel: python export-analytics-to-excel.py \"C:\\path\\to\\тв.xlsx\"")
        sys.exit(1)

    # Индексы колонок (ts=0, page=1, event=2, element=6, durationSec=8)
    page_idx = 1
    event_idx = 2
    element_idx = 6
    duration_idx = 8

    # Агрегация
    event_count = defaultdict(int)
    page_count = defaultdict(int)
    element_count = defaultdict(int)
    page_durations = defaultdict(list)
    sessions = set()

    for r in rows:
        if len(r) <= max(page_idx, event_idx):
            continue
        ev = (r[event_idx] or '').strip() or '(пусто)'
        pg = (r[page_idx] or '').strip() or '(пусто)'
        el = (r[element_idx] or '').strip()
        sid = (r[3] or '').strip()
        try:
            dur = int(r[duration_idx]) if r[duration_idx] else 0
        except (ValueError, TypeError):
            dur = 0

        event_count[ev] += 1
        page_count[pg] += 1
        if ev == 'click' and el:
            element_count[el] += 1
        if sid:
            sessions.add(sid)
        if dur > 0 and ev in ('page_leave', 'visibility_hidden'):
            page_durations[pg].append(dur)

    # Создаём Excel
    out_path = os.path.join(project_root, 'analytics-report.xlsx')
    wb = openpyxl.Workbook()

    # Лист 1: Сырые данные
    ws1 = wb.active
    ws1.title = 'Аналитика'
    ws1.append(COLS)
    for r in rows:
        ws1.append(r)
    for c in range(1, len(COLS) + 1):
        ws1.column_dimensions[get_column_letter(c)].width = 15

    # Лист 2: Сводка
    ws2 = wb.create_sheet('Сводка', 1)
    ws2['A1'] = 'Сводка аналитики'
    ws2['A1'].font = openpyxl.styles.Font(bold=True, size=14)
    row = 3

    # События по типу
    ws2[f'A{row}'] = 'События по типу'
    ws2[f'A{row}'].font = openpyxl.styles.Font(bold=True)
    row += 1
    ws2[f'A{row}'] = 'Тип'
    ws2[f'B{row}'] = 'Кол-во'
    row += 1
    event_data_start = row
    for k, v in sorted(event_count.items(), key=lambda x: -x[1]):
        ws2[f'A{row}'] = k
        ws2[f'B{row}'] = v
        row += 1
    event_data_end = row - 1
    row += 1

    # Просмотры по странице
    ws2[f'A{row}'] = 'Просмотры по странице'
    ws2[f'A{row}'].font = openpyxl.styles.Font(bold=True)
    row += 1
    ws2[f'A{row}'] = 'Страница'
    ws2[f'B{row}'] = 'Кол-во'
    row += 1
    page_data_start = row
    for k, v in sorted(page_count.items(), key=lambda x: -x[1]):
        ws2[f'A{row}'] = k
        ws2[f'B{row}'] = v
        row += 1
    page_data_end = row - 1
    row += 1

    # Топ кликов
    ws2[f'A{row}'] = 'Топ кликов'
    ws2[f'A{row}'].font = openpyxl.styles.Font(bold=True)
    row += 1
    ws2[f'A{row}'] = 'Элемент'
    ws2[f'B{row}'] = 'Кликов'
    row += 1
    for k, v in sorted(element_count.items(), key=lambda x: -x[1])[:15]:
        ws2[f'A{row}'] = k
        ws2[f'B{row}'] = v
        row += 1
    row += 1

    # Среднее время на странице
    ws2[f'A{row}'] = 'Среднее время на странице (сек)'
    ws2[f'A{row}'].font = openpyxl.styles.Font(bold=True)
    row += 1
    ws2[f'A{row}'] = 'Страница'
    ws2[f'B{row}'] = 'Среднее (сек)'
    row += 1
    for pg, durs in sorted(page_durations.items(), key=lambda x: -sum(x[1])/len(x[1]) if x[1] else 0):
        avg = round(sum(durs) / len(durs)) if durs else 0
        ws2[f'A{row}'] = pg
        ws2[f'B{row}'] = avg
        row += 1

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 12

    # Лист 3: Визуальный отчёт — только диаграммы
    ws3 = wb.create_sheet('Визуальный отчёт', 2)
    ws3['A1'] = 'Визуальный отчёт по аналитике'
    ws3['A1'].font = openpyxl.styles.Font(bold=True, size=16)
    ws3.column_dimensions['A'].width = 25

    # Копируем данные для диаграмм (диаграммы должны ссылаться на данные)
    # События по типу — строки 5+
    erow = 5
    ws3['A4'] = 'События по типу'
    ws3['A4'].font = openpyxl.styles.Font(bold=True)
    for k, v in sorted(event_count.items(), key=lambda x: -x[1]):
        ws3[f'A{erow}'] = k
        ws3[f'B{erow}'] = v
        erow += 1
    event_end = erow - 1

    # Просмотры по странице — в колонках D,E
    prow = 5
    ws3['D4'] = 'Просмотры по странице'
    ws3['D4'].font = openpyxl.styles.Font(bold=True)
    for k, v in sorted(page_count.items(), key=lambda x: -x[1]):
        ws3[f'D{prow}'] = k
        ws3[f'E{prow}'] = v
        prow += 1
    page_end = prow - 1

    # Топ кликов — в колонках G,H
    clrow = 5
    ws3['G4'] = 'Топ кликов'
    ws3['G4'].font = openpyxl.styles.Font(bold=True)
    for k, v in sorted(element_count.items(), key=lambda x: -x[1])[:10]:
        ws3[f'G{clrow}'] = (k[:30] + '...') if len(str(k)) > 30 else k
        ws3[f'H{clrow}'] = v
        clrow += 1
    click_end = clrow - 1

    # Диаграмма 1: События по типу (слева вверху)
    if event_end >= 5:
        chart1 = BarChart()
        chart1.type = "col"
        chart1.title = "События по типу"
        chart1.y_axis.title = "Количество"
        data = Reference(ws3, min_col=2, min_row=5, max_row=event_end)
        cats = Reference(ws3, min_col=1, min_row=5, max_row=event_end)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.width = 14
        chart1.height = 10
        ws3.add_chart(chart1, "A20")

    # Диаграмма 2: Просмотры по странице — круговая (справа вверху)
    if page_end >= 5:
        chart2 = PieChart()
        chart2.title = "Просмотры по странице (%)"
        data = Reference(ws3, min_col=5, min_row=5, max_row=page_end)
        cats = Reference(ws3, min_col=4, min_row=5, max_row=page_end)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.width = 14
        chart2.height = 10
        ws3.add_chart(chart2, "J20")

    # Диаграмма 3: Топ кликов — горизонтальная
    if click_end >= 5:
        chart3 = BarChart()
        chart3.type = "bar"
        chart3.title = "Топ кликов"
        chart3.y_axis.title = "Кликов"
        data = Reference(ws3, min_col=8, min_row=5, max_row=click_end)
        cats = Reference(ws3, min_col=7, min_row=5, max_row=click_end)
        chart3.add_data(data, titles_from_data=True)
        chart3.set_categories(cats)
        chart3.width = 14
        chart3.height = 10
        ws3.add_chart(chart3, "A40")

    wb.save(out_path)
    print(f"Готово: {out_path}")
    print(f"  Событий: {len(rows)}, сессий: {len(sessions)}")


if __name__ == '__main__':
    main()
